"""
SPREADSHEET_TOOLS
=================

CSV and Excel spreadsheet tools for the Agentic Loop Framework.

``csv_export``
    Export data to a CSV file. Uses Python's built-in csv module (no
    external dependencies). Output is sandboxed to the agent's runs dir.

``excel_workbook_create``
    Create formatted Excel (.xlsx) spreadsheets with multiple sheets,
    bold headers, and auto-fitted column widths. Requires the ``openpyxl``
    package.

Usage::

    csv_tool = CsvExportTool(agent_dir="/path/to/agent")
    result = csv_tool.execute(
        output_path="report.csv",
        headers=["Name", "Revenue"],
        rows=[["Acme", "50000"], ["Globex", "75000"]],
    )

    xlsx_tool = SpreadsheetCreateTool(agent_dir="/path/to/agent")
    result = xlsx_tool.execute(
        output_path="report.xlsx",
        sheets=[{"name": "Q1", "headers": ["Name", "Revenue"], "rows": [["Acme", "50000"]]}],
    )
"""

import csv
import io
import os
from pathlib import Path
from typing import Dict, List, Optional

from .base import BaseTool, ToolDefinition, ToolParameter, ToolResult


class CsvExportTool(BaseTool):
    """Export data to a CSV file."""

    def __init__(self, agent_dir: str):
        self.agent_dir = Path(agent_dir)

    @property
    def definition(self) -> ToolDefinition:
        return ToolDefinition(
            name="csv_export",
            description=(
                "Export tabular data to a CSV file. Useful for generating reports, "
                "data exports, pipeline logs, or any structured data output."
            ),
            parameters=[
                ToolParameter(
                    name="output_path",
                    type="string",
                    description="Filename relative to agent's runs directory (e.g. 'report.csv')",
                ),
                ToolParameter(
                    name="headers",
                    type="array",
                    description="Column header names",
                    items={"type": "string"},
                ),
                ToolParameter(
                    name="rows",
                    type="array",
                    description="Array of rows, each row is an array of cell values",
                    items={"type": "array", "items": {"type": "string"}},
                ),
                ToolParameter(
                    name="delimiter",
                    type="string",
                    description="Column delimiter character",
                    required=False,
                    default=",",
                ),
            ],
        )

    def execute(
        self,
        output_path: str,
        headers: List[str],
        rows: List[List],
        delimiter: str = ",",
        **kwargs,
    ) -> ToolResult:
        try:
            # Resolve to agent's runs dir
            full_path = self.agent_dir / "runs" / output_path
            full_path.parent.mkdir(parents=True, exist_ok=True)

            with open(full_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, delimiter=delimiter)
                writer.writerow(headers)
                for row in rows:
                    writer.writerow(row)

            size = full_path.stat().st_size
            return ToolResult(
                success=True,
                output=f"CSV saved to {output_path} ({len(rows)} rows, {size} bytes)",
                metadata={
                    "path": str(full_path),
                    "row_count": len(rows),
                    "size_bytes": size,
                },
            )

        except Exception as e:
            return ToolResult(
                success=False,
                output="",
                error=f"CSV export failed: {str(e)}",
            )


class SpreadsheetCreateTool(BaseTool):
    """Create formatted Excel (.xlsx) spreadsheets."""

    def __init__(self, agent_dir: str):
        self.agent_dir = Path(agent_dir)

    @property
    def definition(self) -> ToolDefinition:
        return ToolDefinition(
            name="excel_workbook_create",
            description=(
                "Create a formatted Excel (.xlsx) spreadsheet with multiple sheets, "
                "bold headers, and auto-fitted columns. Use for polished reports, "
                "multi-tab data exports, or client-ready documents."
            ),
            parameters=[
                ToolParameter(
                    name="output_path",
                    type="string",
                    description="Filename relative to agent's runs directory (must end in .xlsx)",
                ),
                ToolParameter(
                    name="sheets",
                    type="array",
                    description=(
                        "Array of sheet definitions. Each sheet has: "
                        "name (string), headers (array of strings), "
                        "rows (array of arrays)"
                    ),
                    items={
                        "type": "object",
                        "properties": {
                            "name": {"type": "string"},
                            "headers": {"type": "array", "items": {"type": "string"}},
                            "rows": {"type": "array", "items": {"type": "array"}},
                        },
                        "required": ["name", "headers", "rows"],
                    },
                ),
                ToolParameter(
                    name="title",
                    type="string",
                    description="Workbook title (metadata)",
                    required=False,
                ),
            ],
        )

    def execute(
        self,
        output_path: str,
        sheets: List[Dict],
        title: str = None,
        **kwargs,
    ) -> ToolResult:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return ToolResult(
                success=False,
                output="",
                error="openpyxl package not installed. Run: pip install openpyxl",
            )

        if not output_path.endswith(".xlsx"):
            return ToolResult(
                success=False,
                output="",
                error="output_path must end in .xlsx",
            )

        try:
            full_path = self.agent_dir / "runs" / output_path
            full_path.parent.mkdir(parents=True, exist_ok=True)

            wb = Workbook()
            # Remove the default sheet
            wb.remove(wb.active)

            if title:
                wb.properties.title = title

            header_font = Font(bold=True)
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            header_text = Font(bold=True, color="FFFFFF")

            total_rows = 0
            for sheet_def in sheets:
                name = sheet_def.get("name", "Sheet")
                headers = sheet_def.get("headers", [])
                rows = sheet_def.get("rows", [])

                ws = wb.create_sheet(title=name)

                # Write headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_text
                    cell.fill = header_fill

                # Write data rows
                for row_idx, row_data in enumerate(rows, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                    total_rows += 1

                # Auto-fit column widths
                for col_idx, header in enumerate(headers, 1):
                    max_len = len(str(header))
                    for row_data in rows:
                        if col_idx - 1 < len(row_data):
                            max_len = max(max_len, len(str(row_data[col_idx - 1])))
                    ws.column_dimensions[
                        ws.cell(row=1, column=col_idx).column_letter
                    ].width = min(max_len + 2, 50)

            wb.save(full_path)
            size = full_path.stat().st_size

            return ToolResult(
                success=True,
                output=(
                    f"Spreadsheet saved to {output_path} "
                    f"({len(sheets)} sheets, {total_rows} rows, {size} bytes)"
                ),
                metadata={
                    "path": str(full_path),
                    "sheet_count": len(sheets),
                    "total_rows": total_rows,
                    "size_bytes": size,
                },
            )

        except Exception as e:
            return ToolResult(
                success=False,
                output="",
                error=f"Spreadsheet creation failed: {str(e)}",
            )
